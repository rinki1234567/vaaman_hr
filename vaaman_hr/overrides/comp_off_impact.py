import frappe
from frappe.utils import flt, getdate, today

from hrms.hr.doctype.leave_application.leave_application import (
	get_leave_allocation_records,
	get_leaves_for_period,
	get_manually_expired_leaves,
	get_remaining_leaves,
)

from vaaman_hr.overrides.comp_off_balance import get_comp_off_balance_on


def hrms_comp_off_balance(employee, date):
	"""Stock HRMS Comp Off display balance (pre-FIFO bug behaviour)."""
	date = getdate(date)
	allocation_records = get_leave_allocation_records(employee, date, "Compensatory Off")
	allocation = allocation_records.get("Compensatory Off")
	if not allocation:
		return 0.0
	leaves_taken = get_leaves_for_period(employee, "Compensatory Off", allocation.from_date, date)
	manually_expired_leaves = get_manually_expired_leaves(
		employee, "Compensatory Off", allocation.from_date, date
	)
	remaining = get_remaining_leaves(allocation, leaves_taken, date, None, manually_expired_leaves)
	return flt(remaining.get("leave_balance"))


def run(date=None):
	d = getdate(date or today())
	employees = frappe.db.sql(
		"""
		SELECT DISTINCT employee
		FROM `tabLeave Ledger Entry`
		WHERE leave_type='Compensatory Off' AND docstatus=1
		""",
		pluck=True,
	)

	neg_old = []
	mismatch = []
	fixed = []
	improved = []

	for emp in employees:
		old = flt(hrms_comp_off_balance(emp, d))
		new = flt(get_comp_off_balance_on(emp, d))
		if abs(old - new) > 0.001:
			mismatch.append((emp, old, new))
		if old < -0.001:
			neg_old.append((emp, old, new))
			if new >= -0.001:
				fixed.append((emp, old, new))
		if new > old + 0.001:
			improved.append((emp, old, new))

	with_leave = frappe.db.sql(
		"""
		SELECT DISTINCT employee FROM `tabLeave Application`
		WHERE leave_type='Compensatory Off' AND docstatus=1 AND status='Approved'
		""",
		pluck=True,
	)

	print(f"As of: {d}")
	print(f"Employees with Comp Off ledger: {len(employees)}")
	print(f"Employees who took Comp Off leave: {len(with_leave)}")
	print("")
	print("=== IMPACT SUMMARY ===")
	print(f"Old method NEGATIVE balance: {len(neg_old)}")
	print(f"Old vs New MISMATCH: {len(mismatch)}")
	print(f"Was negative, FIFO fixed (>=0): {len(fixed)}")
	print(f"Balance IMPROVED under FIFO: {len(improved)}")

	neg_old.sort(key=lambda x: x[1])
	print("")
	print("=== Employees with OLD negative Comp Off ===")
	print(f"{'Employee':<12} {'Name':<40} {'Old':>10} {'New FIFO':>10} {'Diff':>10}")
	for emp, old, new in neg_old:
		name = frappe.db.get_value("Employee", emp, "employee_name") or ""
		print(f"{emp:<12} {name[:40]:<40} {old:10.3f} {new:10.3f} {new - old:10.3f}")

	mismatch.sort(key=lambda x: abs(x[2] - x[1]), reverse=True)
	print("")
	print(f"=== Top 40 mismatches by |diff| (total {len(mismatch)}) ===")
	print(f"{'Employee':<12} {'Name':<40} {'Old':>10} {'New FIFO':>10} {'Diff':>10}")
	for emp, old, new in mismatch[:40]:
		name = frappe.db.get_value("Employee", emp, "employee_name") or ""
		print(f"{emp:<12} {name[:40]:<40} {old:10.3f} {new:10.3f} {new - old:10.3f}")

	return {
		"as_of": str(d),
		"ledger_employees": len(employees),
		"negative_old": len(neg_old),
		"mismatch": len(mismatch),
		"fixed_from_negative": len(fixed),
		"improved": len(improved),
	}


def export_excel(date=None):
	"""Build Excel with mismatch / negative / all Comp Off balances and save to File."""
	from io import BytesIO

	import openpyxl
	from openpyxl.styles import Font

	d = getdate(date or today())
	rows = frappe.db.sql(
		"""
		SELECT DISTINCT lle.employee, e.employee_name, e.branch, e.department, e.status, e.company
		FROM `tabLeave Ledger Entry` lle
		INNER JOIN `tabEmployee` e ON e.name = lle.employee
		WHERE lle.leave_type='Compensatory Off' AND lle.docstatus=1
		ORDER BY lle.employee
		""",
		as_dict=True,
	)

	header = [
		"Employee",
		"Employee Name",
		"Branch",
		"Department",
		"Status",
		"Company",
		"As Of Date",
		"Old Balance (HRMS)",
		"New Balance (FIFO)",
		"Difference",
		"Old Negative?",
		"Fixed From Negative?",
		"Issue Type",
	]
	all_data = [header]
	mismatch_data = [header[:]]
	negative_data = [header[:]]

	seen = set()
	for r in rows:
		if r.employee in seen:
			continue
		seen.add(r.employee)

		old = flt(hrms_comp_off_balance(r.employee, d))
		new = flt(get_comp_off_balance_on(r.employee, d))
		diff = flt(new - old)
		old_neg = "Yes" if old < -0.001 else "No"
		fixed = "Yes" if old < -0.001 and new >= -0.001 else "No"

		if abs(diff) > 0.001:
			if old < -0.001:
				issue = "Negative (Fixed)" if new >= -0.001 else "Negative (Still)"
			elif diff > 0:
				issue = "Understated Balance"
			else:
				issue = "Overstated Balance"
		else:
			issue = "OK"

		row = [
			r.employee,
			r.employee_name or "",
			r.branch or "",
			r.department or "",
			r.status or "",
			r.company or "",
			str(d),
			round(old, 3),
			round(new, 3),
			round(diff, 3),
			old_neg,
			fixed,
			issue,
		]
		all_data.append(row)
		if abs(diff) > 0.001:
			mismatch_data.append(row)
		if old < -0.001:
			negative_data.append(row)

	mismatch_data = [mismatch_data[0]] + sorted(
		mismatch_data[1:], key=lambda x: abs(x[9]), reverse=True
	)
	negative_data = [negative_data[0]] + sorted(negative_data[1:], key=lambda x: x[7])

	def write_sheet(wb, title, data):
		ws = wb.create_sheet(title)
		for i, row in enumerate(data):
			ws.append(row)
			if i == 0:
				for cell in ws[1]:
					cell.font = Font(bold=True)

	wb = openpyxl.Workbook()
	# remove default sheet
	wb.remove(wb.active)
	write_sheet(wb, "Summary_Mismatch", mismatch_data)
	write_sheet(wb, "Old_Negative_Only", negative_data)
	write_sheet(wb, "All_CompOff_Employees", all_data)

	buf = BytesIO()
	wb.save(buf)
	content = buf.getvalue()
	filename = f"Comp_Off_Balance_Impact_{d}.xlsx"

	disk_path = frappe.get_site_path("public", "files", filename)
	with open(disk_path, "wb") as f:
		f.write(content)

	file_doc = frappe.get_doc(
		{
			"doctype": "File",
			"file_name": filename,
			"file_url": f"/files/{filename}",
			"is_private": 0,
		}
	)
	file_doc.insert(ignore_permissions=True)
	frappe.db.commit()

	result = {
		"as_of": str(d),
		"mismatch_count": len(mismatch_data) - 1,
		"negative_count": len(negative_data) - 1,
		"all_count": len(all_data) - 1,
		"file_url": file_doc.file_url,
		"disk_path": disk_path,
		"download_url": f"https://vidhi.vaaman.in{file_doc.file_url}",
	}
	print(result)
	return result
